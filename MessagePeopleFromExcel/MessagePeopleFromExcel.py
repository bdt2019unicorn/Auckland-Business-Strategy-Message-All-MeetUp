from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import openpyxl.styles
import ContactLink


filename = "auckland bussiness strategies.xlsx"
contact_number = 9

mark_contacted_fill_yellow = PatternFill("solid", openpyxl.styles.colors.YELLOW)


workbook = load_workbook(filename)
worksheet = workbook.active 

def firstnoncontactcell(): 
    count = 2 
    global final_fill 
    while True: 
        cell_fill = worksheet["a"+str(count)].fill
        cell_value = worksheet["a"+str(count)].value.strip()
        if(cell_fill.fgColor.rgb == openpyxl.styles.colors.BLACK) or (not isinstance(cell_fill.fgColor.rgb,str)) or (cell_value==""): 
            break
        count+=1
    return count 

count = firstnoncontactcell()
ContactLink.LoginMeetUp()

end_of_loop = count+contact_number
while(count<end_of_loop) and (worksheet["A"+str(count)].value.strip()!=""): 
    cell_value = worksheet["A"+str(count)].value.strip()
    print("Still have "+str(end_of_loop - count))
    ContactLink.MessagePeople(cell_value)
    worksheet["A"+str(count)].fill = mark_contacted_fill_yellow 
    count+=1

workbook.save(filename)
